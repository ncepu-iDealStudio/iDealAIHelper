#!/usr/bin/env python
# -*- coding:utf-8 -*-

# file:promptController.py
# author:song
# datetime:2023/4/25 16:45
# software: PyCharm

"""
this is function  description
"""
import os

import openpyxl
import csv
import json
from datetime import datetime

from sqlalchemy import select

from app.models import Prompt
from db.database import get_async_session_context
from utils.exceptions import InvalidParamsException, DataBaseException
from utils.generate_id import GenerateID


async def get_all_parent_prompt():
    """get all parent prompt
    """
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.prompt_parent_id == 0, Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        return prompts


async def get_prompt_by_prompt_id(prompt_id: str | None):
    """get prompt by prompt id
    """
    async with get_async_session_context() as session:
        if prompt_id is None:
            r = await session.execute(select(Prompt).where(Prompt.is_valid == True, Prompt.prompt_parent_id != 0))
        if prompt_id is not None:
            r = await session.execute(select(Prompt).where(Prompt.prompt_id == prompt_id, Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        return prompts


async def get_prompt_by_parent_id(parent_id: str | None):
    if parent_id is None:
        raise InvalidParamsException("errors.parentIDErroe")
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.prompt_parent_id == parent_id, Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        return prompts

async def delete_prompt_by_parent_id(parent_id: str | None):
    if parent_id is None:
        raise InvalidParamsException("errors.parentIDErroe")
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.prompt_parent_id == parent_id, Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        try:
            for prompt in prompts:
                prompt.is_valid = False
                session.add(prompt)
            result = await session.commit()
        except Exception as e:
            raise DataBaseException(str(e))
        return result


async def update_prompt_by_parent_id(parent_id: str | None, category: str | None):

    if parent_id is None:
        raise InvalidParamsException("errors.parentIDErroe")
    if category is None:
        raise InvalidParamsException("errors.categoryErroe")

    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.prompt_parent_id == parent_id, Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        try:

            for prompt in prompts:
                prompt.category = category

                session.add(prompt)
            result = await session.commit()
        except Exception as e:
            raise DataBaseException(str(e))
        return result


async def add_prompt_by_json_file(target_path: str, parent_id: str | None, category: str | None):
    if parent_id is None:
        parent_id = 0
    if category is None:
        category = 'add_by_json'

    with open(target_path, 'r', encoding='utf-8') as f:
        prompts_data = json.load(f)

    async with get_async_session_context() as session:
        prompt_id_list = []
        prompt_list = []
        try:
            for dict in prompts_data:
                active_time = datetime.utcnow()
                prompt_id = GenerateID.create_random_id()
                new_prompt = Prompt(prompt_parent_id=parent_id, prompt_id=prompt_id,
                                    category=category,
                                    title=dict['act'], prompt=dict['prompt'], create_time=active_time)
                prompt_list.append(new_prompt)
                prompt_id_list.append(prompt_id)
            session.add_all(prompt_list)
            await session.commit()
            result = {
                'prompt_id_list': prompt_id_list,
                'prompt_parent_id': 0
            }
        except Exception as e:
            raise DataBaseException(str(e))
        return result


async def add_prompt_by_csv_file(target_path: str, parent_id: str | None, category: str | None):
    if parent_id is None:
        parent_id = 0
    if category is None:
        category = 'add_by_csv'

    prompts_data = []
    with open(target_path, 'r', encoding='utf-8') as f:
        csv_reader = csv.reader(f)
        header = next(csv_reader)
        for row in csv_reader:
            temp_dict = {
                'act': row[0],
                'prompt': row[1]
            }
            prompts_data.append(temp_dict)
    async with get_async_session_context() as session:
        prompt_id_list = []
        prompt_list = []
        try:
            for dict in prompts_data:
                active_time = datetime.utcnow()
                prompt_id = GenerateID.create_random_id()
                new_prompt = Prompt(prompt_parent_id=parent_id, prompt_id=prompt_id,
                                    category=category,
                                    title=dict['act'], prompt=dict['prompt'], create_time=active_time)
                prompt_list.append(new_prompt)
                prompt_id_list.append(prompt_id)
            session.add_all(prompt_list)
            await session.commit()
            result = {
                'prompt_id_list': prompt_id_list,
                'prompt_parent_id': 0
            }
        except Exception as e:
            raise DataBaseException(str(e))
        return result


async def get_prompt_by_prompt_category(category: str):
    """
    根据分类查看是否有该分类，没有则创建一个分类，返回该分类的id
    """
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.category == category, Prompt.is_valid == True,
                                                       Prompt.prompt_parent_id == 0))
        prompt = r.scalars().one_or_none()
        # 判断分类是否存在
        if prompt is not None:
            parent_id = prompt.prompt_id
        if prompt is None:
            try:
                active_time = datetime.utcnow()
                prompt_id = GenerateID.create_random_id()
                new_prompt = Prompt(prompt_parent_id=0, prompt_id=prompt_id,
                                    category=category,
                                    title=category, prompt=category, create_time=active_time)
                parent_id = prompt_id
                session.add(new_prompt)
                await session.commit()
            except Exception as e:
                raise DataBaseException(str(e))

        return parent_id


async def get_prompt_by_title(title: str):
    """
    判断该prompt是否存在 若不存在返回true
    """
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.title == title, Prompt.is_valid == True,
                                                       Prompt.prompt_parent_id != 0))
        prompt = r.scalars().one_or_none()
        if prompt is not None:
            return True
        else:
            return False


async def add_prompts_by_xlsx_file(target_path: str, filename: str):
    """
    通过.xlsx文件添加prompt
    """
    workbook = openpyxl.load_workbook(target_path)  # 加载Excel文件
    worksheet = workbook.active  # 获取工作表
    prompt_data = []  # 文件中的prompt
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        temp_dict = {}
        temp_dict['category'] = str(row[0])
        temp_dict['title'] = str(row[1])
        temp_dict['prompt'] = str(row[2])
        temp_dict['prompt_id'] = GenerateID.create_random_id()
        prompt_data.append(temp_dict)

    # prompt和父级直间的映射
    son_by_parent_dict = {}
    # 添加失败的列表
    fail_list = []
    for p in prompt_data:
        son_by_parent_dict[p['prompt_id']] = await get_prompt_by_prompt_category(p['category'])
    async with get_async_session_context() as session:
        try:
            for p in prompt_data:
                title = p['title']

                # 判断prompt是已经否存在
                flag = await get_prompt_by_title(title)
                if flag:
                    fail_list.append(p)
                    continue

                # 判断是否为一级分类
                category = p['category']
                if title == category:
                    continue

                active_time = datetime.utcnow()
                prompt_id = p['prompt_id']
                prompt_parent_id = son_by_parent_dict[prompt_id]
                prompt = p['prompt']
                new_prompt = Prompt(prompt_parent_id=prompt_parent_id, prompt_id=prompt_id,
                                    category=category,
                                    title=title, prompt=prompt, create_time=active_time)
                session.add(new_prompt)
            await session.commit()
        except Exception as e:
            raise DataBaseException(str(e))

    return {"prompt_data": prompt_data, "fail_list": fail_list}


async def get_prompt_to_file(target_path):
    """
    获取数据库中所有prompt文件
    """
    async with get_async_session_context() as session:
        r = await session.execute(select(Prompt).where(Prompt.is_valid == True))
        prompts = r.scalars().all()
        if prompts is None:
            raise InvalidParamsException("errors.promptsNotFound")
        prompts_list = []
        # 将数据库数据写入列表
        for prompt in prompts:
            temp_dict = {}
            temp_dict['category'] = prompt.category
            temp_dict['title'] = prompt.title
            temp_dict['prompt'] = prompt.prompt
            prompts_list.append(temp_dict)
        try:
            filename = "all_prompts.xlsx"
            file_path = os.path.join(target_path, filename)

            if os.path.isfile(file_path):
                # 文件已存在，删除原来的数据并写入新数据
                workbook = openpyxl.load_workbook(file_path)  # 加载Excel文件
                worksheet = workbook.active  # 获取活动工作表

                # 更新数据
                for row_index, row_data in enumerate(prompts_list, start=2):
                    worksheet.cell(row=row_index, column=1, value=row_data['category'])
                    worksheet.cell(row=row_index, column=2, value=row_data['title'])
                    worksheet.cell(row=row_index, column=3, value=row_data['prompt'])
                # 设置列宽
                worksheet.column_dimensions["C"].width = 200
                workbook.save(file_path)  # 将工作簿保存到文件系统中
            else:
                # 文件不存在，创建一个新的工作簿并写入数据
                workbook = openpyxl.Workbook()  # 创建一个新的工作簿
                worksheet = workbook.active  # 获取工作簿的活动工作表

                # 写入标题数据
                worksheet["A1"] = "分类"
                worksheet["B1"] = "标题"
                worksheet["C1"] = "提示词"

                for row_index, row_data in enumerate(prompts_list, start=2):
                    worksheet.cell(row=row_index, column=1, value=row_data['category'])
                    worksheet.cell(row=row_index, column=2, value=row_data['title'])
                    worksheet.cell(row=row_index, column=3, value=row_data['prompt'])

                worksheet.column_dimensions["C"].width = 400
                workbook.save(file_path)  # 将工作簿保存到文件系统中

        except Exception as e:
            raise DataBaseException(str(e))

        return file_path
